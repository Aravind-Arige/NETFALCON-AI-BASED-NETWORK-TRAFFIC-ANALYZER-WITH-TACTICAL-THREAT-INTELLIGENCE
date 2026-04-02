import time
from flask import Flask, jsonify, request, render_template
from analyzer.engine import (
    start_live_capture,
    stop_live_capture,
    analyze_pcap,
    metrics,
    threats,
    protocols,
    get_capture_status,
    get_interfaces_list,
    inject_test_threat
)
from analyzer.simulator import simulator_engine

from analyzer.database import db
from analyzer.mitre import group_threats_into_campaigns
import os
import re

from flask_socketio import SocketIO, emit

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
capture_thread = None
dos_sim_active = False

# ── AI Anomaly Detection Engine ──────────────────────────────────
from anomaly_engine.engine import AnomalyDetectionEngine
anomaly_engine = AnomalyDetectionEngine(alert_cooldown=60)
anomaly_engine.load_models()   # no-op if no saved model yet

# ── Smart Alert Dispatcher ───────────────────────────────────────
try:
    from alerts.dispatcher import AlertDispatcher
    alert_dispatcher: AlertDispatcher | None = AlertDispatcher()
    anomaly_engine.on_anomaly(
        lambda alert: alert_dispatcher.dispatch(alert, alert.get("raw_metrics", {})) if alert_dispatcher else None
    )
    print("[App] Alert dispatcher connected to anomaly engine.")
except Exception as _ae:
    alert_dispatcher = None
    print(f"[App] Alert dispatcher not loaded: {_ae}")


# ── Background thread ────────────────────────────────────────────
def background_thread():
    loop_count = 0
    last_threat_idx = 0
    
    while True:
        time.sleep(1)
        loop_count += 1
        state = get_capture_status()

        unmitigated = sum(1 for t in threats if not t.get("mitigated"))
        health_score = 100
        health_score -= min(60, unmitigated * 15)

        latency_val = metrics.latency()
        if latency_val > 300:   health_score -= 20
        elif latency_val > 150: health_score -= 10
        elif latency_val > 100: health_score -= 5

        jitter_val = metrics.jitter()
        if jitter_val > 80:     health_score -= 10
        elif jitter_val > 30:   health_score -= 5

        loss_val = metrics.packet_loss()
        if loss_val > 50:       health_score -= 20
        elif loss_val > 10:     health_score -= 10
        elif loss_val > 0:      health_score -= 5

        err_pct = metrics.error_percentage()
        if err_pct > 10:        health_score -= 10
        elif err_pct > 5:       health_score -= 5

        health_score = max(0, min(100, health_score))

        # Feed metrics into the AI engine every cycle
        bw_kbps = metrics.bandwidth()
        anomaly_metrics = {
            "bandwidth_in":    bw_kbps,
            "bandwidth_out":   bw_kbps * 0.3,
            "packet_loss":     metrics.packet_loss(),
            "latency":         metrics.latency(),
            "jitter":          metrics.jitter(),
            "error_rate":      metrics.error_rate(),
            "active_flows":    state.get("flow_count", 0),
            "packets_per_sec": state.get("packet_rate", 0),
            "unique_ips":      len(state.get("suspicious_ips", [])),
            "timestamp":       time.time(),
        }
        anomaly_result = anomaly_engine.ingest(anomaly_metrics)



        data = {
            "running": state["running"],
            "capturing": (time.time() - state["last_packet_time"] < 5) if state["last_packet_time"] else False,
            "packet_count": metrics.packet_count,
            "bandwidth": metrics.bandwidth(),
            "network_speed": metrics.speed_bps(),
            "speed_text": metrics.formatted_speed(),
            "latency": metrics.latency(),
            "jitter": metrics.jitter(),
            "threats": len(threats),
            "threat_list": threats[-20:],
            "packet_loss": metrics.packet_loss(),
            "protocols": protocols,
            "packet_rate": state.get("packet_rate", 0),
            "flow_count": state.get("flow_count", 0),
            "error_rate": metrics.error_rate(),
            "error_percentage": metrics.error_percentage(),
            "active_flows": state.get("flows", []),
            "dos_sim_active": dos_sim_active,
            "top_talkers": state.get("top_talkers", []),
            "suspicious_ips": state.get("suspicious_ips", []),
            "health_score": health_score,
            "current_interface": state.get("current_interface"),
            "anomaly": anomaly_result,

        }
        
        # ── Database Persistence (batched every ~10 seconds) ──
        if loop_count % 10 == 0 and state["running"]:
            db.log_metrics({**metrics.__dict__, "health_score": health_score})
            db.log_top_talkers(state.get("top_talkers", []))
            
        # ── Log New Threats Instantly ──
        current_threat_len = len(threats)
        if current_threat_len > last_threat_idx:
            for t in threats[last_threat_idx:current_threat_len]:
                db.log_threat(t)
            last_threat_idx = current_threat_len
            
        # Nightly Purge check (rough approximation)
        if loop_count % 86400 == 0:
            db.purge_old_data(days=7)

        socketio.emit('status_update', data)


@app.before_request
def start_background_task():
    pass


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/dashboard")
def dashboard():
    return render_template("index.html")


@app.route("/interfaces")
def interfaces():
    return jsonify(get_interfaces_list())


@app.route("/start", methods=["POST"])
def start():
    data = request.json or {}
    interface_name = data.get("interface")
    if interface_name == "auto":
        interface_name = None
    start_live_capture(interface_name)
    return jsonify({"status": "running", "interface": interface_name})


@app.route("/stop", methods=["POST"])
def stop():
    stop_live_capture()
    return jsonify({"status": "stopped"})


@app.route("/stats")
def stats():
    state = get_capture_status()
    return jsonify({
        "packet_count": metrics.packet_count,
        "bandwidth": metrics.bandwidth(),
        "network_speed": metrics.speed_bps(),
        "speed_text": metrics.formatted_speed(),
        "latency": metrics.latency(),
        "jitter": metrics.jitter(),
        "threats": len(threats),
        "threat_list": threats[-20:],
        "packet_loss": metrics.packet_loss(),
        "protocols": protocols,
        "packet_rate": state.get("packet_rate", 0),
        "flow_count": state.get("flow_count", 0),
        "error_rate": metrics.error_rate(),
        "error_percentage": metrics.error_percentage()
    })


@app.route("/status")
def status():
    state = get_capture_status()
    now = time.time()
    active = False
    last_packet_time = state["last_packet_time"]
    if last_packet_time:
        active = (now - last_packet_time) < 5
    return jsonify({"running": state["running"], "capturing": active})


@app.route("/upload_pcap", methods=["POST"])
def upload_pcap():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file part"}), 400
        file = request.files["file"]
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        base_dir = os.path.dirname(os.path.abspath(__file__))
        upload_dir = os.path.join(base_dir, "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, filename)
        file.save(path)
        print(f"[DEBUG] Analyzing PCAP: {path}")
        analyze_pcap(path)
        return jsonify({"status": "pcap analyzed", "filename": filename})
    except Exception as e:
        print(f"[ERROR] PCAP Upload/Analysis failed: {e}")
        return jsonify({"error": str(e)}), 500



# ── Historical Data Query API ────────────────────────────────────
@app.route("/api/history/query", methods=["POST"])
def history_query():
    data = request.json or {}
    start_time = data.get("start_time")
    end_time = data.get("end_time")
    
    # Default to last 1 hour if not specified
    if not end_time: end_time = time.time()
    if not start_time: start_time = end_time - 3600
        
    ip_filter = data.get("ip_filter")
    res = db.query_history(start_time, end_time, ip_filter)
    return jsonify(res)

@app.route("/api/killchain")
def api_killchain():
    try:
        campaigns = group_threats_into_campaigns(threats)
        # Safe-serialize: convert any numpy/non-standard types to native Python
        import json
        def safe_default(obj):
            if hasattr(obj, 'item'):
                return obj.item()  # numpy scalar -> python scalar
            return str(obj)
        safe_json = json.loads(json.dumps({"campaigns": campaigns}, default=safe_default))
        return jsonify(safe_json)
    except Exception as e:
        print(f"[KillChain] Error: {e}")
        return jsonify({"campaigns": [], "error": str(e)})

@app.route("/api/simulate/start", methods=["POST"])
def simulate_start():
    data = request.json or {}
    sim_id = data.get("sim_id", "default")
    sim_type = data.get("type", "port_scan")
    intensity = int(data.get("intensity", 5))
    safe_mode = data.get("safe_mode", True)
    
    if simulator_engine.start(sim_id, sim_type, intensity, safe_mode):
        return jsonify({"status": "started", "sim_id": sim_id})
    return jsonify({"error": "Simulation already running"}), 400

@app.route("/api/simulate/stop", methods=["POST"])
def simulate_stop():
    data = request.json or {}
    sim_id = data.get("sim_id", "default")
    if simulator_engine.stop(sim_id):
        return jsonify({"status": "stopped", "sim_id": sim_id})
    return jsonify({"error": "Simulation not found"}), 404

@app.route("/api/simulate/reset", methods=["POST"])
def simulate_reset():
    """Emergency: force-clear all simulation state regardless of running flag."""
    simulator_engine.reset()
    return jsonify({"status": "reset"})

@app.route("/api/simulate/report")
def simulate_report():
    status = simulator_engine.get_status()
    # Simple evaluation of engine detecting these packets (in real world this queries the DB)
    detected_alerts = sum(1 for t in threats[-100:] if t.get("metadata", {}).get("type") == "Developer Test" or t.get("mitre", {}).get("stage", 0) > 0)
    
    return jsonify({
        "status": status,
        "report": {
            "total_threats_logged": detected_alerts,
            "engine_active": True
        }
    })
# ─────────────────────────────────────────────────────────────────

# ── IP Blocking ──────────────────────────────────────────────────
from analyzer.firewall import block_ip_firewall, unblock_ip_firewall, get_blocked_ips as retrieve_blocked_ips


@app.route("/block_ip", methods=["POST"])
def block_ip():
    data = request.json or {}
    ip = data.get("ip", "").strip()
    if not ip:
        return jsonify({"error": "No IP provided"}), 400
    import re
    if not re.match(r'^\d{1,3}(\.\d{1,3}){3}$', ip):
        return jsonify({"error": "Invalid IP address"}), 400
    success = block_ip_firewall(ip)
    if success:
        return jsonify({"status": "blocked", "ip": ip})
    return jsonify({"status": "tracked", "ip": ip, "note": "Added to session block list"})


@app.route("/unblock_ip", methods=["POST"])
def unblock_ip():
    data = request.json or {}
    ip = data.get("ip", "").strip()
    if not ip:
        return jsonify({"error": "No IP provided"}), 400
    success = unblock_ip_firewall(ip)
    if success:
        return jsonify({"status": "unblocked", "ip": ip})
    return jsonify({"status": "removed_from_session", "ip": ip})


@app.route("/blocked_ips")
def get_blocked_ips():
    return jsonify({"blocked": retrieve_blocked_ips()})


@app.route("/mitigate_threat", methods=["POST"])
def mitigate_threat():
    data = request.json or {}
    index = data.get("index")
    if index is None or not isinstance(index, int) or index < 0 or index >= len(threats):
        return jsonify({"error": "Invalid threat index"}), 400
    threat = threats[index]
    if threat.get("mitigated"):
        return jsonify({"status": "already_mitigated"})
    metadata = threat.get("metadata", {})
    direct_action = metadata.get("direct_action")
    if not direct_action:
        return jsonify({"error": "No direct mitigation available"}), 400
    action_type = direct_action.get("type")
    if action_type == "block_ip":
        ip = direct_action.get("ip")
        success = block_ip_firewall(ip)
        threat["mitigated"] = True
        if success:
            return jsonify({"status": "mitigated", "ip": ip})
        return jsonify({"status": "tracked", "ip": ip, "note": "Added to session block list"})
    return jsonify({"error": "Unknown action type"}), 400


# ── Anomaly Engine REST API ──────────────────────────────────────

@app.route("/api/anomaly/current")
def get_anomaly_current():
    return jsonify({
        "score":             anomaly_engine.current_score,
        "label":             anomaly_engine.current_label,
        "is_learning":       anomaly_engine.is_learning,
        "learning_progress": anomaly_engine.learning_progress,
    })


@app.route("/api/anomaly/history")
def get_anomaly_history():
    return jsonify(anomaly_engine.get_score_history())


@app.route("/api/anomaly/alerts")
def get_anomaly_alerts():
    return jsonify(anomaly_engine.get_alert_history())


@app.route("/api/anomaly/retrain", methods=["POST"])
def retrain_anomaly():
    anomaly_engine.retrain()
    return jsonify({"status": "retraining_started"})


@app.route("/api/anomaly/save", methods=["POST"])
def save_anomaly_model():
    anomaly_engine.save_models()
    return jsonify({"status": "model_saved"})


# ── Report generation (unchanged) ────────────────────────────────

from typing import Any, Dict
def generate_report_data() -> Dict[str, Any]:
    summ = metrics.summary()
    return {
        'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
        'packet_count': summ['packet_count'],
        'total_bytes': round(metrics.total_bytes / 1024, 2),
        'bandwidth': summ['bandwidth'],
        'latency': summ['latency'],
        'jitter': summ['jitter'],
        'protocols': summ['protocols'],
        'threats': threats[-100:] if threats else []
    }


def generate_txt_report():
    data = generate_report_data()
    report = []
    report.append("="*50)
    report.append("       NETWORK ANALYSIS SESSION REPORT")
    report.append("="*50)
    report.append(f"Generated at: {data['timestamp']}")
    report.append("-" * 50)
    report.append(" [CORE METRICS]")
    report.append(f" Total Packets Captured: {data['packet_count']}")
    report.append(f" Total Data Volume:      {data['total_bytes']} KB")
    report.append(f" Avg Bandwidth:          {data['bandwidth']} KB/s")
    report.append(f" Latency (Avg):          {data['latency']} ms")
    report.append(f" Jitter (Avg):           {data['jitter']} ms")
    report.append("-" * 50)
    report.append(" [PROTOCOL DISTRIBUTION]")
    if not data['protocols']:
        report.append(" No protocol data recorded.")
    else:
        for proto, count in data['protocols'].items():
            report.append(f" {proto:<15} : {count} packets")
    report.append("-" * 50)
    report.append(" [SECURITY THREAT LOG]")
    if not data['threats']:
        report.append(" No security threats detected during this session.")
    else:
        report.append(f" Detected Threats: {len(data['threats'])}")
        report.append("")
        for t in data['threats']:
            report.append(f" [{t['timestamp']}] {t['alert']}")
            report.append(f"             Source: {t['src']} -> Dest: {t['dst']}")
    report.append("="*50)
    report.append("            END OF ANALYSIS REPORT")
    report.append("="*50)
    return "\n".join(report)


def generate_pdf_report():
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    import io

    data = generate_report_data()
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                 fontSize=24, textColor=colors.HexColor('#00e5ff'),
                                 spaceAfter=30, alignment=1)
    elements.append(Paragraph("NETWORK ANALYSIS SESSION REPORT", title_style))
    elements.append(Spacer(1, 0.3*inch))

    ts_style = ParagraphStyle('Timestamp', parent=styles['Normal'],
                              fontSize=10, textColor=colors.grey)
    elements.append(Paragraph(f"<b>Generated at:</b> {data['timestamp']}", ts_style))
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph("<b>CORE METRICS</b>", styles['Heading2']))
    metrics_data = [
        ['Metric', 'Value'],
        ['Total Packets', str(data['packet_count'])],
        ['Total Data', f"{data['total_bytes']} KB"],
        ['Avg Bandwidth', f"{data['bandwidth']} KB/s"],
        ['Latency', f"{data['latency']} ms"],
        ['Jitter', f"{data['jitter']} ms"],
    ]
    tbl = Table(metrics_data, colWidths=[3*inch, 2*inch])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d333b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph("<b>PROTOCOL DISTRIBUTION</b>", styles['Heading2']))
    if data['protocols']:
        pd = [['Protocol', 'Packets']]
        for proto, count in data['protocols'].items():
            pd.append([proto, str(count)])
        pt = Table(pd, colWidths=[3*inch, 2*inch])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d333b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(pt)
    else:
        elements.append(Paragraph("No protocol data.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph("<b>SECURITY THREAT LOG</b>", styles['Heading2']))
    if not data['threats']:
        elements.append(Paragraph("No threats detected.", styles['Normal']))
    else:
        elements.append(Paragraph(f"<b>Detected: {len(data['threats'])}</b>", styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        td = [['Timestamp', 'Alert', 'Source -> Dest']]
        for t in data['threats'][:50]:
            td.append([
                t['timestamp'][:19],
                (t['alert'][:50] + '...') if len(t['alert']) > 50 else t['alert'],
                f"{t['src']} -> {t['dst']}"
            ])
        tt = Table(td, colWidths=[1.2*inch, 2*inch, 2*inch])
        tt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(tt)

    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


def generate_excel_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io

    data = generate_report_data()
    wb = Workbook()
    ws = wb.active  # type: ignore
    ws.title = "Network Analysis"  # type: ignore

    ws['A1'] = "NETWORK ANALYSIS SESSION REPORT"  # type: ignore
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='00E5FF')  # type: ignore
    ws.merge_cells('A1:D1')  # type: ignore
    ws['A2'] = f"Generated at: {data['timestamp']}"  # type: ignore
    ws.merge_cells('A2:D2')  # type: ignore

    hfill = PatternFill(start_color="2d333b", end_color="2d333b", fill_type="solid")
    hfont = Font(name='Arial', size=10, bold=True, color="FFFFFF")

    row = 4
    ws[f'A{row}'] = "CORE METRICS"  # type: ignore
    ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)  # type: ignore
    row += 1
    for col, h in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=row, column=col)  # type: ignore
        c.value = h; c.fill = hfill; c.font = hfont
    row += 1
    for m_name, m_val in [
        ("Total Packets", str(data['packet_count'])),
        ("Total Data", f"{data['total_bytes']} KB"),
        ("Avg Bandwidth", f"{data['bandwidth']} KB/s"),
        ("Latency", f"{data['latency']} ms"),
        ("Jitter", f"{data['jitter']} ms"),
    ]:
        ws[f'A{row}'] = m_name  # type: ignore
        ws[f'B{row}'] = m_val  # type: ignore
        row += 1

    row += 2
    ws[f'A{row}'] = "PROTOCOL DISTRIBUTION"  # type: ignore
    ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)  # type: ignore
    row += 1
    for col, h in enumerate(["Protocol", "Packets"], 1):
        c = ws.cell(row=row, column=col)  # type: ignore
        c.value = h; c.fill = hfill; c.font = hfont
    row += 1
    if data['protocols']:
        for proto, count in data['protocols'].items():
            ws[f'A{row}'] = proto; ws[f'B{row}'] = count; row += 1  # type: ignore
    else:
        ws[f'A{row}'] = "No data"; row += 1  # type: ignore

    row += 2
    ws[f'A{row}'] = "SECURITY THREAT LOG"  # type: ignore
    ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)  # type: ignore
    row += 1
    if not data['threats']:
        ws[f'A{row}'] = "No threats detected."  # type: ignore
    else:
        ws[f'A{row}'] = f"Detected: {len(data['threats'])}"  # type: ignore
        row += 1
        for col, h in enumerate(["Timestamp", "Alert", "Source", "Destination"], 1):
            c = ws.cell(row=row, column=col)  # type: ignore
            c.value = h; c.fill = hfill; c.font = hfont
        row += 1
        for t in data['threats']:
            ws[f'A{row}'] = t['timestamp']  # type: ignore
            ws[f'B{row}'] = t['alert']  # type: ignore
            ws[f'C{row}'] = t['src']  # type: ignore
            ws[f'D{row}'] = t['dst']  # type: ignore
            row += 1

    ws.column_dimensions['A'].width = 20  # type: ignore
    ws.column_dimensions['B'].width = 25  # type: ignore
    ws.column_dimensions['C'].width = 20  # type: ignore
    ws.column_dimensions['D'].width = 20  # type: ignore

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/download_report")
def download_report():
    from flask import make_response
    output = make_response(generate_txt_report())
    output.headers["Content-Disposition"] = "attachment; filename=network_analysis_report.txt"
    output.headers["Content-type"] = "text/plain"
    return output


@app.route("/download_report/<fmt>")
def download_report_format(fmt):
    from flask import make_response
    if fmt == 'txt':
        content = generate_txt_report()
        output = make_response(content)
        output.headers["Content-Disposition"] = "attachment; filename=network_analysis_report.txt"
        output.headers["Content-type"] = "text/plain"
    elif fmt == 'pdf':
        content = generate_pdf_report()
        output = make_response(content)
        output.headers["Content-Disposition"] = "attachment; filename=network_analysis_report.pdf"
        output.headers["Content-type"] = "application/pdf"
    elif fmt == 'excel':
        content = generate_excel_report()
        output = make_response(content)
        output.headers["Content-Disposition"] = "attachment; filename=network_analysis_report.xlsx"
        output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        return jsonify({"error": "Invalid format"}), 400
    return output


if __name__ == "__main__":
    socketio.start_background_task(target=background_thread)
    primary = 5001
    try:
        socketio.run(app, debug=True, port=primary)
    except OSError as e:
        print(f"[WARN] Port {primary} unavailable ({e}), switching to ephemeral port")
        socketio.run(app, debug=True, port=0)
